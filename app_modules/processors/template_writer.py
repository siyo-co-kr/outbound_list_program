"""resources/outbound_template.xlsx 의 양식 그대로 결과 엑셀을 생성한다.

서식을 코드로 재현하지 않고 템플릿 파일 자체를 열어서 값만 채우므로,
글꼴·테두리·머리글 색·열 너비·작성예시 블록이 원본과 100% 동일하게 유지된다.
"""

import os
import re
import sys
from copy import copy

import openpyxl

TEMPLATE_NAME = "outbound_template.xlsx"

# 템플릿 A1:C1 머리글과 동일한 순서
TEMPLATE_COLUMNS = ["차트번호", "환자 이름", "휴대폰번호"]

_FIRST_DATA_ROW = 2


def resource_path(*parts):
    """PyInstaller 번들 실행과 소스 실행 양쪽에서 통하는 리소스 절대 경로"""
    base = getattr(sys, "_MEIPASS", None)
    if base is None:
        base = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    return os.path.join(base, *parts)


def _load_template():
    path = resource_path("resources", TEMPLATE_NAME)
    if not os.path.exists(path):
        raise FileNotFoundError(f"양식 파일을 찾을 수 없습니다: {path}")
    return openpyxl.load_workbook(path)


def _to_text(value):
    """텍스트 서식(@) 셀에 넣을 문자열로 변환. 12345.0 / '12345.0' -> '12345'"""
    if value is None or value != value:                  # None 또는 NaN
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text = str(value).strip()
    if text.lower() == "nan":
        return None

    # 엑셀이 숫자로 저장한 셀은 문자열로 읽어도 '123.0' 형태가 될 수 있다
    # (.xls는 xlrd가 모든 숫자를 float으로 돌려준다). 차트번호가 '123.0'으로
    # 저장되지 않도록, 값 전체가 정수+소수점0일 때만 떼어낸다.
    text = re.sub(r'^(\d+)\.0+$', r'\1', text)

    return text or None


def _copy_style(ws, src_row, src_col, dst_row, dst_col):
    ws.cell(row=dst_row, column=dst_col)._style = copy(
        ws.cell(row=src_row, column=src_col)._style
    )


def _ensure_styled_rows(ws, column_count, needed_rows):
    """데이터 행이 템플릿에 미리 서식이 잡힌 행 수를 넘으면 서식을 이어서 복제"""
    last_row = _FIRST_DATA_ROW + needed_rows - 1
    for row in range(ws.max_row + 1, last_row + 1):
        for col in range(1, column_count + 1):
            _copy_style(ws, _FIRST_DATA_ROW, col, row, col)


def _write_rows(ws, rows):
    for offset, record in enumerate(rows):
        row = _FIRST_DATA_ROW + offset
        for col, value in enumerate(record, start=1):
            ws.cell(row=row, column=col).value = _to_text(value)
    return len(rows)


def save_outbound_list(records, save_path):
    """아웃바운드 리스트 필터 결과를 템플릿 양식 그대로 저장한다.

    records - {컬럼명: 값} 목록. TEMPLATE_COLUMNS 순서로 채워 넣는다.
    """
    missing = sorted({c for r in records for c in TEMPLATE_COLUMNS if c not in r})
    if missing:
        raise ValueError(f"양식에 필요한 컬럼이 없습니다: {', '.join(missing)}")

    wb = _load_template()
    ws = wb.worksheets[0]

    rows = [[record[column] for column in TEMPLATE_COLUMNS] for record in records]
    _ensure_styled_rows(ws, len(TEMPLATE_COLUMNS), len(rows))
    count = _write_rows(ws, rows)

    wb.save(save_path)
    return count
